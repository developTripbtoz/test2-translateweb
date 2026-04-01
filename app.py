from flask import Flask, request, send_file, jsonify
import openpyxl
import xlrd
import urllib.request
import urllib.parse
import json
import io
import time
import os

app = Flask(__name__)

def translate_text(text):
    """Google Translate 비공식 API로 중국어 → 한국어 번역"""
    if not text or not str(text).strip():
        return ''
    s = str(text).strip()
    encoded = urllib.parse.quote(s)
    url = f'https://translate.googleapis.com/translate_a/single?client=gtx&sl=zh-CN&tl=ko&dt=t&q={encoded}'
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
            return ''.join(part[0] for part in data[0] if part[0])
    except Exception:
        return s

HTML_PAGE = '''<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>중국어 → 한국어 엑셀 번역기</title>
  <style>
    *{margin:0;padding:0;box-sizing:border-box}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f4f8;min-height:100vh;display:flex;align-items:center;justify-content:center}
    .container{background:#fff;border-radius:16px;padding:40px;max-width:520px;width:90%;box-shadow:0 4px 24px rgba(0,0,0,.1);text-align:center}
    h1{font-size:1.6rem;margin-bottom:8px;color:#1a1a2e}
    .subtitle{color:#666;margin-bottom:28px;font-size:.95rem}
    .upload-area{border:2px dashed #cbd5e1;border-radius:12px;padding:40px 20px;cursor:pointer;transition:all .3s;margin-bottom:20px}
    .upload-area:hover{border-color:#3b82f6;background:#eff6ff}
    .upload-area.dragover{border-color:#3b82f6;background:#dbeafe}
    .upload-icon{font-size:2.5rem;margin-bottom:12px}
    .upload-text{color:#64748b;font-size:.9rem}
    .file-name{color:#3b82f6;font-weight:600;margin-top:8px}
    input[type=file]{display:none}
    button{background:#3b82f6;color:#fff;border:none;padding:14px 32px;border-radius:10px;font-size:1rem;cursor:pointer;transition:background .3s;width:100%}
    button:hover{background:#2563eb}
    button:disabled{background:#94a3b8;cursor:not-allowed}
    .progress{display:none;margin-top:20px}
    .progress-bar{background:#e2e8f0;border-radius:8px;height:8px;overflow:hidden}
    .progress-fill{background:#3b82f6;height:100%;width:0%;transition:width .3s;border-radius:8px}
    .progress-text{color:#64748b;font-size:.85rem;margin-top:8px}
    .result{display:none;margin-top:20px;padding:16px;background:#f0fdf4;border-radius:10px}
    .result a{color:#16a34a;font-weight:600;text-decoration:none}
    .error{display:none;margin-top:20px;padding:16px;background:#fef2f2;border-radius:10px;color:#dc2626}
  </style>
</head>
<body>
  <div class="container">
    <h1>🇨🇳 → 🇰🇷 엑셀 번역기</h1>
    <p class="subtitle">중국어 엑셀 파일을 업로드하면 한국어로 번역해드립니다</p>
    <form id="uploadForm" enctype="multipart/form-data">
      <div class="upload-area" id="dropZone">
        <div class="upload-icon">📄</div>
        <div class="upload-text">클릭하거나 파일을 드래그하세요</div>
        <div class="upload-text" style="font-size:.8rem;margin-top:4px">.xlsx, .xls 파일 지원 (최대 10MB)</div>
        <div class="file-name" id="fileName"></div>
      </div>
      <input type="file" id="fileInput" name="file" accept=".xlsx,.xls">
      <button type="submit" id="submitBtn" disabled>번역 시작</button>
    </form>
    <div class="progress" id="progress">
      <div class="progress-bar"><div class="progress-fill" id="progressFill"></div></div>
      <div class="progress-text" id="progressText">번역 중...</div>
    </div>
    <div class="result" id="result">✅ 번역 완료! <a id="downloadLink" href="#">번역된 파일 다운로드</a></div>
    <div class="error" id="error"></div>
  </div>
  <script>
    const dropZone=document.getElementById('dropZone'),fileInput=document.getElementById('fileInput'),fileName=document.getElementById('fileName'),submitBtn=document.getElementById('submitBtn'),form=document.getElementById('uploadForm'),progress=document.getElementById('progress'),progressFill=document.getElementById('progressFill'),progressText=document.getElementById('progressText'),result=document.getElementById('result'),error=document.getElementById('error');
    dropZone.addEventListener('click',()=>fileInput.click());
    dropZone.addEventListener('dragover',e=>{e.preventDefault();dropZone.classList.add('dragover')});
    dropZone.addEventListener('dragleave',()=>dropZone.classList.remove('dragover'));
    dropZone.addEventListener('drop',e=>{e.preventDefault();dropZone.classList.remove('dragover');if(e.dataTransfer.files.length){fileInput.files=e.dataTransfer.files;updateFileName()}});
    fileInput.addEventListener('change',updateFileName);
    function updateFileName(){if(fileInput.files.length){fileName.textContent=fileInput.files[0].name;submitBtn.disabled=false}}
    form.addEventListener('submit',async e=>{
      e.preventDefault();if(!fileInput.files.length)return;
      submitBtn.disabled=true;progress.style.display='block';result.style.display='none';error.style.display='none';
      progressFill.style.width='20%';progressText.textContent='파일 업로드 중...';
      const formData=new FormData();formData.append('file',fileInput.files[0]);
      try{
        progressFill.style.width='50%';progressText.textContent='번역 중... (셀 수에 따라 시간이 걸릴 수 있습니다)';
        const res=await fetch('/translate',{method:'POST',body:formData});
        if(!res.ok){const err=await res.json();throw new Error(err.error||'번역 실패')}
        progressFill.style.width='90%';progressText.textContent='파일 생성 중...';
        const blob=await res.blob();const url=URL.createObjectURL(blob);
        const dl=document.getElementById('downloadLink');dl.href=url;dl.download='번역_'+fileInput.files[0].name.replace(/\\.xls$/,'.xlsx');
        progressFill.style.width='100%';progressText.textContent='완료!';result.style.display='block';
      }catch(err){error.textContent='❌ '+err.message;error.style.display='block'}
      finally{submitBtn.disabled=false}
    });
  </script>
</body>
</html>'''


@app.route('/')
def index():
    return HTML_PAGE


@app.route('/translate', methods=['POST'])
def translate():
    if 'file' not in request.files:
        return jsonify(error='파일이 없습니다.'), 400

    file = request.files['file']
    filename = file.filename.lower()

    if not filename.endswith(('.xlsx', '.xls')):
        return jsonify(error='엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'), 400

    try:
        file_bytes = file.read()
        if len(file_bytes) > 10 * 1024 * 1024:
            return jsonify(error='파일 크기가 10MB를 초과합니다.'), 400

        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)

        if filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                translated_name = translate_text(sheet_name)[:31]
                out_ws = out_wb.create_sheet(title=translated_name)
                for row in ws.iter_rows():
                    for cell in row:
                        val = cell.value
                        if val and isinstance(val, str) and val.strip():
                            out_ws.cell(row=cell.row, column=cell.column, value=translate_text(val))
                            time.sleep(0.05)
                        else:
                            out_ws.cell(row=cell.row, column=cell.column, value=val)
        else:
            wb = xlrd.open_workbook(file_contents=file_bytes)
            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                translated_name = translate_text(ws.name)[:31]
                out_ws = out_wb.create_sheet(title=translated_name)
                for r in range(ws.nrows):
                    for c in range(ws.ncols):
                        val = ws.cell_value(r, c)
                        if val and isinstance(val, str) and val.strip():
                            out_ws.cell(row=r+1, column=c+1, value=translate_text(val))
                            time.sleep(0.05)
                        else:
                            out_ws.cell(row=r+1, column=c+1, value=val)

        output = io.BytesIO()
        out_wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='translated.xlsx'
        )

    except Exception as e:
        return jsonify(error=f'번역 중 오류가 발생했습니다: {str(e)}'), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
